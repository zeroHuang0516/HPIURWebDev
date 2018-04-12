from django.shortcuts import render

# Create your views here.
from datetime import datetime
from django.shortcuts import render
from django.http import HttpResponse
from django.views.generic import View
from main.models import Employee
from main.models import Unit, Dash_request

from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
from django.core.files import File
from django.http import HttpResponseRedirect

import csv
import json
import codecs
import xlrd
from datetime import datetime
from xlrd import xldate_as_tuple

import smtplib

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
from io import StringIO
import xlsxwriter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from mongoengine.django.auth import User 

def login_view(request):
    #connect('testdb')
    from django.contrib.auth import login
    from mongoengine.django.auth import User
    from mongoengine.queryset import DoesNotExist
    from django.contrib import messages
    error_msg = ''
    try:
        user = User.objects.get(username=request.POST['username'])
        if user.check_password(request.POST['password']):
            user.backend = 'mongoengine.django.auth.MongoEngineBackend'
            login(request, user)
            request.session.set_expiry(60 * 60 * 1) # 1 hour timeout
            if user['is_superuser']:
            	return HttpResponseRedirect('/index/')
            else:
            	return HttpResponseRedirect('/uindex/')
        else:
            error_msg = u"Incorrect login name or password !"
            error_msg = json.dumps(error_msg)
            return render(request, 'hello_world.html',{'error_msg':error_msg})
    except DoesNotExist:
        error_msg = u"Incorrect login name or password !"
        error_msg = json.dumps(error_msg)
        return render(request, 'hello_world.html',{'error_msg':error_msg})
    error_msg = json.dumps(error_msg)
    return render(request, 'hello_world.html', {})

def logout(request):
    from django.contrib.auth import logout
    logout(request)
    return HttpResponseRedirect('/login/')

def createuser(request): 
    User.create_user(request.POST['username'],request.POST['password'])
    return HttpResponseRedirect('/login/')


def index(request):
	#from django.contrib.auth import logout
	#logout(request.user)
	if request.user.is_authenticated() and request.user['is_superuser']:
		return render(request, 'indexPage.html')
	else:
		return HttpResponseRedirect('/login/')

def uindex(request):
	#from django.contrib.auth import logout
	#logout(request.user)
	if request.user.is_authenticated() and not request.user['is_superuser']:
		return render(request, 'usrIndexPage.html')


def login_page(request):
	if request.user.is_authenticated() and request.user['is_superuser']:
	    return render(request, 'indexPage.html')
	elif request.user.is_authenticated() and not request.user['is_superuser']:
	    return render(request, 'usrIndexPage.html')
	else:
	    return render(request, 'hello_world.html')

def check_in_units(request):
	return render(request, 'checkIn.html')

def units_borrow(request):
	
	if request.method == 'POST' and 'search_box' in request.POST : # If the form is submitted
		search_data = []
		search_query = request.POST.get('search_box', None)		
		search_query = str(search_query)
		print(search_query)
		if search_query !="":
			for u in Unit.objects.all():
				if u["move_in_date"].lower() == search_query.lower() \
							or u["owner"].lower() == search_query.lower() \
							or u["core_team"].lower() == search_query.lower() \
							or u["platform"].lower() == search_query.lower() \
							or u["phase"].lower() == search_query.lower() \
							or u["sku"].lower() == search_query.lower() \
							or u["cat_no"].lower() == search_query.lower() \
							or u["s_n"].lower() == search_query.lower() \
							or u["comment"].lower() == search_query.lower() \
							or u["status"].lower() == search_query.lower():
					r = {}
					r['move_in_date'] = (u["move_in_date"])
					r['owner'] = (u["owner"])
					r['core_team'] = (u["core_team"])
					r['platform'] = (u["platform"])
					r['phase'] = (u["phase"])
					r['sku'] = (u["sku"])
					r['cat_no'] = (u["cat_no"])
					r['s_n'] = (u["s_n"])
					r['comment'] = (u["comment"])
					r['status'] = (u["status"])
					search_data.append(r)
		print(len(search_data))
		search_data = json.dumps(search_data)
		print(search_data)
		request.session['search_data'] = search_data
		return HttpResponseRedirect('/search/result/')
		#url = reverse('')
		#return HttpResponseRedirect('/search/result/',{'search_data': search_data})
		#return render(request,'search_result.html',{'search_data': search_data})
	

	elif request.method == 'POST' and 'borrow_cat' in request.POST:
	   
	   borrow_person = request.POST.get('borrow_person', None)
	   borrow_cat = request.POST.get('borrow_cat', None)
	   borrow_date = request.POST.get('borrow_date', None)
	   borrow_purpose = request.POST.get('borrow_purpose', None)
	   return_status = {'status': ""}

	   if borrow_cat !="":
		   for u in Unit.objects.all():
		   		if u["cat_no"] == borrow_cat:
		   			if u["status"] == "In Pool":
		   				u["status"] = "Borrowed"
		   				existing_unit = u
		   				existing_unit["owner"] = borrow_person
		   				existing_unit["borrow_date"] = borrow_date
		   				existing_unit["borrow_purpose"] = borrow_purpose
		   				print(borrow_person)
		   				existing_unit.save()
		   				return_status = { 'status': "Send Borrow Request Success!!!"}
		   			else:
		   				return_status = {'status': "The Unit has been borrowed!!!"}
		   			break
		   		else:
		   			return_status = {'status': "The Unit cannot be found!!!"}
	   #print(borrow_cat)
	   #print(return_status)	
	   return_status = json.dumps(return_status)
	   if not request.POST._mutable:
   			request.POST._mutable = True
	   del request.POST['borrow_person']
	   del request.POST['borrow_cat']
	   del request.POST['borrow_date']
	   del request.POST['borrow_purpose']
	   #print(request.POST)
	   return render(request,'borrow.html',{'return_status': return_status})
	return render(request,'borrow.html')

def search_results(request):
	search_data = []
	if request.session.has_key('search_data'):
		search_data = request.session.get('search_data')
		del request.session['search_data']
	return render(request,'search_result.html',{'search_data': search_data})

def units_list(request):
	if request.method == 'POST' and 'borrow_cat' in request.POST :   
	   print(request.POST.get('borrow_cat'))
	   existing_unit = Unit.objects(cat_no= request.POST.get('borrow_cat'))[0]
	   existing_unit["status"] = "Borrowed"
	   existing_unit["owner"] = request.POST.get('borrow_person')
	   existing_unit["borrow_date"] = request.POST.get('borrow_date')
	   existing_unit["borrow_purpose"] = request.POST.get('borrow_purpose')	   
	   existing_unit.save()
	elif request.method == 'POST' and 'return_no' in request.POST :
	   print(request.POST.get('return_no')) 
	   existing_unit = Unit.objects(cat_no= request.POST.get('return_no'))[0]
	   existing_unit["status"] = "In Pool"
	   existing_unit["owner"] = "TDC"
	   existing_unit["borrow_date"] = ""
	   existing_unit["borrow_purpose"] = ""   
	   existing_unit.save()

	data=[]
	for u in Unit.objects.all():
		r = {}
		r['move_in_date'] = (u["move_in_date"])
		r['owner'] = (u["owner"])
		r['core_team'] = (u["core_team"])
		r['platform'] = (u["platform"])
		r['phase'] = (u["phase"])
		r['sku'] = (u["sku"])
		r['cat_no'] = (u["cat_no"])
		r['s_n'] = (u["s_n"])
		r['comment'] = (u["comment"])
		r['status'] = (u["status"])
		r['borrow_date'] = (u["borrow_date"])
		r['borrow_purpose'] = (u["borrow_purpose"])
		data.append(r)
	data = json.dumps(data)
	#print(data)
	return render(request, 'units_list.html',{'list_data':data})

class myView(View):
    def post(self,request):
        csvfile = request.FILES['csv_file']
        workbook = xlrd.open_workbook(file_contents=csvfile.read())
        #csvfile = csvfile.read().decode('utf-8')
        sheet = workbook.sheet_by_index(0)
        offset = 1
        data = []
        for i, row in enumerate(range(sheet.nrows)):
        	if i <= offset:
        		continue
        	r = []
        	for j, col in enumerate(range(sheet.ncols)):
        		cell = sheet.cell_value(i, j)
        		# move_in_date
        		if j == 0 and cell !='':
        			date = datetime(*xldate_as_tuple(cell,0))
        			cell = date.strftime('%Y/%m/%d')
        		r.append(cell)
        	data.append(r)

        print(len(data))
        print((data[30]))
        print(type(data[30][0]))

        for each_row in range(len(data)):
        	new_unit = Unit( move_in_date = str(data[each_row][0]),
        					 	owner = str(data[each_row][1]),
        					 	core_team = str(data[each_row][2]),
        					 	platform = str(data[each_row][3]),
        					 	phase = str(data[each_row][4]),
        					 	sku = str(data[each_row][5]),
        					 	cat_no = str(data[each_row][6]),
        					 	s_n = str(data[each_row][7]),
        					 	comment = str(data[each_row][8]),
        						status = "In Pool",
        						borrow_date = "",
        						borrow_purpose ="",
        					)
        	if (len(Unit.objects(cat_no= str(data[each_row][6]))) == 0):
        		new_unit.save()

        
        return HttpResponseRedirect('/index/')

def send_mail_via_com(receive_email_adr, platform, requirement):

	outlook_user = 'erica.huang@hp.com'
	outlook_pwd = '@hyukoh09830927'

	smtpserver = smtplib.SMTP('smtp-mail.outlook.com', 587)
	smtpserver.ehlo()
	smtpserver.starttls()
	smtpserver.ehlo()

	smtpserver.login(outlook_user, outlook_pwd)

	msg = MIMEMultipart('alternative')
	msg['Subject'] = 'IUR Pool Reminder'
	msg['From'] = outlook_user
	msg['To'] = receive_email_adr

	html = """\
	<html>
	  <head></head>
	  <body>
	    <h3>Hi!<br>
	       The """+ requirement+""" task of """+ platform+""" has been finsihed!<br>
	       Please come to IUR Pool to reclaim it.
	    </h3>
	    <h4>*** This is an automatically generated email, please do not reply ***</h4>
	  </body>
	</html>
	"""
	part2 = MIMEText(html, 'html')
	msg.attach(part2)

	
	 
	smtpserver.sendmail(outlook_user, receive_email_adr, msg.as_string())
	 

	smtpserver.quit()
	print("send email")

def send_email(request):
	if request.method == 'POST':
		existing_request = Dash_request.objects(fill_in_datetime= request.POST.get('fill_in_datetime'))[0]
		existing_request["press_send_btn_time"] = request.POST.get('fill_in_datetime')
		existing_request.save()
		#print(request.POST.get('press_send_btn_time'))
		send_mail_via_com(request.POST.get('email_adr'),request.POST.get('platform'),request.POST.get('requirement'))

	data = []
	for u in Dash_request.objects.all():
		r = {}
		r['status'] = (u["status"])
		r['fill_in_datetime'] = (u["fill_in_datetime"])
		r['applicant_email'] = (u["applicant_email"])
		r['department'] = (u["department"])
		r['requirement'] = (u['requirement'])
		r['project_name'] = (u["project_name"])
		r['phase'] = (u["phase"])
		r['bios_version'] = (u["bios_version"])
		r['os'] = (u["os"])
		r['language'] = (u["language"])
		r['issue'] = (u["issue"])
		data.append(r)
	data = json.dumps(data)

	return render(request, 'b_d_request_list.html',{'list_data':data})

def b_d_request(request):
	if request.user.is_authenticated():
		if request.method == 'POST' and 'fill_in_datetime' in request.POST:
		   status = ''
		   fill_in_datetime = request.POST.get('fill_in_datetime', None)
		   press_send_btn_time = ''
		   applicant_email = request.POST.get('applicant_email', None)
		   department = request.POST.get('department', None)
		   project_name = ''
		   requirement = request.POST.get('requirement', None)
		   phase = ''
		   bios_version = ''
		   OS = ''
		   language = ''
		   password = ''
		   issue = ''
		   if requirement == 'Update BIOS':
	   			project_name = request.POST.get('bios_project_name', None)
	   			phase = request.POST.get('bios_phase', None)
	   			bios_version = request.POST.get('bios_bios_version', None)
	   			password = request.POST.get('bios_set_password', None)
	   			issue = request.POST.get('bios_issue', None)
		   elif requirement == 'Dash Image':
	   			project_name = request.POST.get('dash_project_name', None)
	   			phase = request.POST.get('dash_phase', None)
	   			OS = request.POST.get('dash_OS', None)
	   			language = request.POST.get('dash_language', None)
	   			password = request.POST.get('dash_set_password', None)
	   			issue = request.POST.get('dash_issue', None)
		   else:
		   		project_name = request.POST.get('both_project_name', None)
		   		phase = request.POST.get('both_phase', None)
		   		bios_version = request.POST.get('both_bios_version', None)
		   		OS = request.POST.get('both_OS', None)
		   		language = request.POST.get('both_language', None)
		   		password = request.POST.get('set_password', None)
		   		issue = request.POST.get('both_issue', None)
		   
		   return_status = {'status': ""}
		   
		   new_dash_request = Dash_request(
		   							status = status,
		   							fill_in_datetime = fill_in_datetime,
		   							press_send_btn_time = press_send_btn_time,
	        					 	applicant_email = applicant_email,
	        					 	department = department,
	        					 	requirement = requirement,
	        					 	project_name = project_name,
	        					 	phase = phase,
	        					 	bios_version = bios_version,
	        					 	os = OS,
	        					 	language = language,
	        					 	issue = issue,
	        					 	use = "test",
	        					 	password = password,
	        					)

		   if(len(Dash_request.objects(fill_in_datetime= str(fill_in_datetime))) == 0):
	   			new_dash_request.save()
		   
		   return_status = json.dumps(return_status)
		   if not request.POST._mutable:
	   			request.POST._mutable = True
		   #del request.POST['borrow_person']
		   #del request.POST['borrow_cat']
		   #del request.POST['borrow_date']
		   #del request.POST['borrow_purpose']
		   #print(request.POST)
		   return HttpResponseRedirect('/b_d_request/success')
		return render(request,'b_d_request.html')
	else:
		return HttpResponseRedirect('/login/')

def b_d_request_success(request):
	return render(request, 'b_d_request_success.html')

def set_sheet_col_width(ws):
   dim_dist = {}
   for row in ws.rows:
      for cell in row:
         if cell.value:
            dim_dist[cell.column] = max((dim_dist.get(cell.column, 0), len(str(cell.value))+7))
            
   for col, value in dim_dist.items():
      ws.column_dimensions[col].width=value
   return ws

def report_down_load(request):
	print(request.method)
	if request.method == 'POST':
		from_datetime = request.POST.get('from_datetime')
		to_datetime = request.POST.get('to_datetime') 
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=report.xlsx'
	wb = openpyxl.Workbook()
	ws = wb.get_active_sheet()
	ws.title = "Records"
	count_ws = wb.create_sheet(title="Statistics")
	header_list = ['Fill in Datetime','Applicant\'s Emails','Product Name','Requirement','BIOS','OS','Language','Department']
	
	for row in range(1,2):
		for col in range(1,9):
			ws.cell(row=row, column=col).value = header_list[col-1]
			ws.cell(row=row,column=col).fill = PatternFill(fgColor="d2e9ff", fill_type = "solid")
			ws.cell(row=row,column=col).border = Border(left=Side(style='thin',color='0072e3'),
														right=Side(style='thin',color='0072e3'), 
														top=Side(style='thin',color='0072e3'), 
														bottom=Side(style='thin',color='0072e3'))

	data = []
	statistics=[0,0,0] #both/bios/dash
	for u in Dash_request.objects.all():
		r = []
		if((u["requirement"])) == 'Dash Image & Update BIOS':
			statistics[0] += 1
		elif ((u["requirement"])) == 'Update BIOS':
			statistics[1] += 1
		elif ((u["requirement"])) == 'Dash Image':
			statistics[2] += 1

		r.append((u["fill_in_datetime"]))
		r.append((u["applicant_email"]))
		r.append((u["project_name"]))
		r.append((u['requirement']))
		r.append((u["bios_version"]))
		r.append((u["os"]))
		r.append((u["language"]))
		r.append((u["department"]))
		data.append(r)

	for row in range(2,len(data)+2):
		for col in range(1,9):
			ws.cell(row=row, column=col).value = data[row-2][col-1]
			if row%2 != 0:
				ws.cell(row=row,column=col).fill = PatternFill(fgColor="f0f0f0", fill_type = "solid")
	
	ws = set_sheet_col_width(ws)

	# for sheet Count
	count_ws.cell(row=1,column=1).value = 'Service Status-Period from' + from_datetime + ' to '+ to_datetime
	count_ws.cell(row=2,column=1).value = 'Items'
	count_ws.cell(row=2,column=2).value = 'no.'
	count_ws.cell(row=2,column=3).value = 'Note'
	count_ws.cell(row=3,column=1).value = '#Request of Image refresh with BIOS update'
	count_ws.cell(row=4,column=1).value = '#Request of BIOS update only'
	count_ws.cell(row=5,column=1).value = '#Request of Image refresh only'
	for j in range(3,6):
		count_ws.cell(row=j,column=2).value = str(statistics[j-3]) +' units'
		count_ws.cell(row=j,column=3).value = '100 of request delivered in 4hrs'

	count_ws = set_sheet_col_width(count_ws)
	count_ws.merge_cells('A1:C1')
	count_ws.cell(row=1,column=1).alignment = Alignment(horizontal='center')
	count_ws.cell(row=1,column=1).fill = PatternFill(fgColor="FFD306", fill_type = "solid")
	count_ws.cell(row=1,column=1).border = Border(left=Side(style='thin',color='5b5b5b'),
														right=Side(style='thin',color='5b5b5b'), 
														top=Side(style='thin',color='5b5b5b'), 
														bottom=Side(style='thin',color='5b5b5b'))
	for p in range(2,6):
		for k in range(1,4):
			if(p==2):
				count_ws.cell(row=2,column=k).fill = PatternFill(fgColor="FFE665", fill_type = "solid")
				count_ws.cell(row=2,column=k).border = Border(left=Side(style='thin',color='5b5b5b'),
														right=Side(style='thin',color='5b5b5b'), 
														top=Side(style='thin',color='5b5b5b'), 
														bottom=Side(style='thin',color='5b5b5b'))
			else:
				count_ws.cell(row=p,column=k).fill = PatternFill(fgColor="f0f0f0", fill_type = "solid")
				count_ws.cell(row=p,column=k).border = Border(left=Side(style='thin',color='5b5b5b'),
														right=Side(style='thin',color='5b5b5b'), 
														top=Side(style='thin',color='5b5b5b'), 
														bottom=Side(style='thin',color='5b5b5b'))


	wb.save(response)
	return response


def b_d_request_list(request):
	if request.user.is_authenticated() and request.user['is_superuser']:
		if request.method == 'POST':
			status = request.POST.get('status')
			existing_request = Dash_request.objects(fill_in_datetime= request.POST.get('fill_in_datetime'))[0]
			#print(existing_request[0]["status"])
			existing_request["status"] = status
			existing_request.save()

		data = []
		for u in Dash_request.objects.all():
			r = {}
			r['status'] = (u["status"])
			r['fill_in_datetime'] = (u["fill_in_datetime"])
			r['applicant_email'] = (u["applicant_email"])
			r['department'] = (u["department"])
			r['requirement'] = (u['requirement'])
			r['project_name'] = (u["project_name"])
			r['phase'] = (u["phase"])
			r['bios_version'] = (u["bios_version"])
			r['os'] = (u["os"])
			r['language'] = (u["language"])
			r['issue'] = (u["issue"])
			data.append(r)
		data = json.dumps(data)

		return render(request, 'b_d_request_list.html',{'list_data':data})
	else:
		return HttpResponseRedirect('/login/')
